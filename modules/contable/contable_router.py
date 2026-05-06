"""
modules/contable/contable_router.py
"""
from __future__ import annotations
import io
from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from auth.internal_auth import require_internal_auth
from modules.caja.comisiones_store import calcular as calcular_comision
from db.supabase_client import get_gastos, get_pagos_mes

router = APIRouter(prefix="/api/contable", tags=["Contable"])

TIPOS_LABELS  = {"particular":"Particular","control_costo":"Control con costo","control_gratuito":"Control gratuito","sobrecupo":"Sobrecupo","cortesia":"Cortesía","kinesiologia":"Kinesiología"}
GRUPOS_LABELS = {"fijos":"Gastos Fijos","variables":"Gastos Variables","cuentas":"Cuentas"}


def _get_profesionales_scope(auth: dict) -> list | None:
    """
    Retorna lista de professional IDs permitidos según scope.
    None = sin restricción (scope ica o admin).
    """
    scope = (auth.get("role") or {}).get("scope", "ica")
    if scope == "ica":
        return None  # ve todo
    # externo → solo su professional
    professional = auth.get("professional")
    if professional and professional != "system":
        return [professional]
    return []


def _resumen_mes(mes: str, profesionales_permitidos: list | None = None) -> dict:
    pagos  = get_pagos_mes(mes)
    gastos = get_gastos().get(mes, {})

    # Filtrar pagos por scope
    if profesionales_permitidos is not None:
        pagos = [p for p in pagos if p.get("professional") in profesionales_permitidos]

    ingresos_total=0; anulados_total=0; pago_profesionales=0
    ingresos_detalle=[]; anulados_detalle=[]; profesionales_detalle={}
    for p in pagos:
        monto=p.get("monto",0); prof=p["professional"]
        if p.get("anulado"):
            anulados_total+=monto
            anulados_detalle.append({"fecha":p["fecha"],"time":p["time"],"rut":p.get("rut",""),
                "tipo":TIPOS_LABELS.get(p.get("tipo_atencion",""),p.get("tipo_atencion","")),"monto":-monto,
                "motivo":p.get("anulacion_motivo",""),"professional":prof})
        else:
            comision=calcular_comision(prof,monto)
            ingresos_total+=monto; pago_profesionales+=comision["neto"]
            ingresos_detalle.append({"fecha":p["fecha"],"time":p["time"],"rut":p.get("rut",""),
                "tipo":TIPOS_LABELS.get(p.get("tipo_atencion",""),p.get("tipo_atencion","")),
                "monto":monto,"retencion":comision["retencion"],"neto":comision["neto"],
                "porcentaje":comision["porcentaje"],"metodo":p.get("metodo_pago") or "gratuito",
                "es_gratuito":p.get("es_gratuito",False),"professional":prof})
            if prof not in profesionales_detalle:
                profesionales_detalle[prof]={"monto":0,"retencion":0,"neto":0,"pagos":0}
            profesionales_detalle[prof]["monto"]+=monto
            profesionales_detalle[prof]["retencion"]+=comision["retencion"]
            profesionales_detalle[prof]["neto"]+=comision["neto"]
            profesionales_detalle[prof]["pagos"]+=1

    # Gastos solo para scope ica — externo no ve gastos del centro
    if profesionales_permitidos is not None:
        gastos = {}

    gastos_total=0; gastos_detalle=[]; gastos_por_grupo={}
    for grupo, items in gastos.items():
        grupo_total=0
        for gasto in items:
            monto=gasto.get("monto",0); gastos_total+=monto; grupo_total+=monto
            gastos_detalle.append({"id":gasto["id"],"grupo":GRUPOS_LABELS.get(grupo,grupo),
                "categoria":gasto["categoria"],"descripcion":gasto.get("descripcion",""),
                "monto":-monto,"created_at":gasto.get("created_at","")})
        gastos_por_grupo[GRUPOS_LABELS.get(grupo,grupo)]=grupo_total
    gastos_por_grupo["Pago Profesionales"]=pago_profesionales
    return {
        "mes":mes,"ingresos_total":ingresos_total,"anulados_total":anulados_total,
        "pago_profesionales":pago_profesionales,"gastos_total":gastos_total,
        "utilidad_neta":ingresos_total-anulados_total-pago_profesionales-gastos_total,
        "ingresos":sorted(ingresos_detalle,key=lambda x:(x["fecha"],x["time"])),
        "anulados":sorted(anulados_detalle,key=lambda x:(x["fecha"],x["time"])),
        "gastos":sorted(gastos_detalle,key=lambda x:x["created_at"]),
        "gastos_por_grupo":gastos_por_grupo,"profesionales_detalle":profesionales_detalle,
    }


@router.get("/resumen/{mes}")
def get_resumen(mes: str, auth=Depends(require_internal_auth)):
    permitidos = _get_profesionales_scope(auth)
    return _resumen_mes(mes, permitidos)


@router.get("/exportar/{mes}")
def exportar_excel(mes: str, auth=Depends(require_internal_auth)):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    permitidos = _get_profesionales_scope(auth)
    data=_resumen_mes(mes, permitidos); wb=Workbook()
    NAVY="0D1B2A";BLUE="1E40AF";GREEN="166534";RED="991B1B";YELLOW="854D0E";PURPLE="5B21B6"
    BG_GRAY="F1F5F9";BG_GREEN="F0FDF4";BG_RED="FEF2F2";BG_BLUE="EFF6FF";BG_YELLOW="FEFCE8";BG_PURPLE="F5F3FF"
    def hf(c=None): return Font(name="Arial",bold=True,color=c or "FFFFFF",size=11)
    def cf(bold=False,color="0F172A",size=10): return Font(name="Arial",bold=bold,color=color,size=size)
    def fill(c): return PatternFill("solid",start_color=c,fgColor=c)
    def brd():
        t=Side(style="thin",color="E2E8F0")
        return Border(left=t,right=t,top=t,bottom=t)
    def ctr(): return Alignment(horizontal="center",vertical="center")
    ws=wb.active; ws.title="Resumen"; ws.sheet_view.showGridLines=False
    ws.merge_cells("A1:F1"); ws["A1"]=f"ICA — Resumen Contable {mes}"
    ws["A1"].font=Font(name="Arial",bold=True,size=14,color=NAVY); ws["A1"].alignment=ctr()
    kpis=[("Ingresos brutos",data["ingresos_total"],BG_GREEN,GREEN),
          ("Anulaciones",-data["anulados_total"],BG_RED,RED),
          ("Pago profesionales",-data["pago_profesionales"],BG_PURPLE,PURPLE),
          ("Gastos operacionales",-data["gastos_total"],BG_YELLOW,YELLOW),
          ("Utilidad neta",data["utilidad_neta"],BG_BLUE,BLUE)]
    for i,(label,valor,bg,color) in enumerate(kpis):
        col=i+1
        c1=ws.cell(row=5,column=col,value=label); c1.font=Font(name="Arial",size=9,bold=True,color="64748B")
        c1.alignment=Alignment(horizontal="center"); c1.fill=fill(bg); c1.border=brd()
        c2=ws.cell(row=6,column=col,value=valor); c2.font=Font(name="Arial",size=14,bold=True,color=color)
        c2.alignment=ctr(); c2.fill=fill(bg); c2.number_format="$#,##0;($#,##0);"-""; c2.border=brd()
    for col,w in zip("ABCDE",[32,10,18,18,18]): ws.column_dimensions[col].width=w
    ws2=wb.create_sheet("Ingresos"); ws2.sheet_view.showGridLines=False
    ws2.merge_cells("A1:I1"); ws2["A1"]=f"Ingresos — {mes}"
    ws2["A1"].font=Font(name="Arial",bold=True,size=13,color=NAVY); ws2["A1"].alignment=ctr()
    for j,h in enumerate(["Fecha","Hora","RUT","Tipo atención","Profesional","Método","Bruto ($)","Retención ($)","Neto prof. ($)"]):
        c=ws2.cell(row=2,column=j+1,value=h); c.font=hf(); c.fill=fill(BLUE); c.alignment=ctr(); c.border=brd()
    for i,ing in enumerate(data["ingresos"]):
        rn=i+3; bg="FFFFFF" if i%2==0 else BG_GRAY
        for j,v in enumerate([ing["fecha"],ing["time"],ing["rut"],ing["tipo"],ing["professional"],ing["metodo"],ing["monto"],ing["retencion"],ing["neto"]]):
            c=ws2.cell(row=rn,column=j+1,value=v); c.font=cf(); c.fill=fill(bg); c.border=brd()
            if j==6: c.number_format="$#,##0;($#,##0);"-""; c.font=cf(bold=True,color=GREEN)
            if j==7: c.number_format="$#,##0;($#,##0);"-""; c.font=cf(bold=True,color=BLUE)
            if j==8: c.number_format="$#,##0;($#,##0);"-""; c.font=cf(bold=True,color=PURPLE)
    for col,w in zip("ABCDEFGHI",[12,8,14,20,16,14,14,14,14]): ws2.column_dimensions[col].width=w
    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(buf,media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition":f"attachment; filename=contable_{mes}.xlsx"})
                
