"""
modules/rrhh/excel_liquidacion.py
Genera Excel resumen de liquidaciones del mes.
"""

from __future__ import annotations

import io


def generar_excel(data: dict, buffer: io.BytesIO) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = f"Remuneraciones {data['mes']}"
    ws.sheet_view.showGridLines = False

    NAVY = "0D1B2A"
    GRAY = "F1F5F9"

    def hf(c="FFFFFF", bold=True, sz=10):
        return Font(name="Arial", bold=bold, color=c, size=sz)

    def fill(c):
        return PatternFill("solid", start_color=c, fgColor=c)

    def brd():
        t = Side(style="thin", color="E2E8F0")
        return Border(left=t, right=t, top=t, bottom=t)

    def ctr():
        return Alignment(horizontal="center", vertical="center")

    # Título
    ws.merge_cells("A1:P1")
    ws["A1"] = f"Instituto de Cirugía Articular — Remuneraciones {data['mes']}"
    ws["A1"].font      = Font(name="Arial", bold=True, size=14, color=NAVY)
    ws["A1"].alignment = ctr()
    ws.row_dimensions[1].height = 32

    # KPIs filas 3-4
    kpis = [
        ("Trabajadores",    data["trabajadores"],          "F1F5F9"),
        ("Total líquido",   data["total_liquidos"],         "F0FDF4"),
        ("Total desc.",     data["total_descuentos"],       "FEF2F2"),
        ("Costo empresa",   data["total_costo_empresa"],    "EFF6FF"),
    ]
    for i, (label, val, bg) in enumerate(kpis):
        col = i * 2 + 1
        ws.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col+1)
        ws.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col+1)
        c1 = ws.cell(row=3, column=col, value=label)
        c1.font = Font(name="Arial", size=9, bold=True, color="64748B")
        c1.alignment = ctr(); c1.fill = fill(bg); c1.border = brd()
        c2 = ws.cell(row=4, column=col, value=val)
        c2.font = Font(name="Arial", size=14, bold=True, color=NAVY)
        c2.alignment = ctr(); c2.fill = fill(bg); c2.border = brd()
        c2.number_format = "$#,##0;($#,##0);\"-\""
    ws.row_dimensions[3].height = 18
    ws.row_dimensions[4].height = 30

    # Headers
    headers = [
        "Nombre", "RUT", "Cargo", "Contrato", "AFP", "Salud",
        "Sueldo Base", "Bono Col.", "Bono Mov.",
        "AFP Desc.", "Salud Desc.", "AFC", "Imp. Único",
        "Total Desc.", "Líquido", "Costo Empresa"
    ]
    for j, h in enumerate(headers):
        c = ws.cell(row=6, column=j+1, value=h)
        c.font = hf(); c.fill = fill(NAVY); c.alignment = ctr(); c.border = brd()
    ws.row_dimensions[6].height = 22

    # Datos
    for i, liq in enumerate(data["liquidaciones"]):
        rn = i + 7
        bg = "FFFFFF" if i % 2 == 0 else GRAY

        bonos = liq.get("bonos", [])
        b_col = next((b["monto"] for b in bonos if "colac" in b["nombre"].lower()), 0)
        b_mov = next((b["monto"] for b in bonos if "movil" in b["nombre"].lower()), 0)

        if liq.get("es_honorarios"):
            vals = [
                liq["nombre"], liq.get("rut",""), liq.get("cargo",""),
                "Honorarios", "-", "-",
                liq["sueldo_base"], 0, 0, 0, 0, 0, 0,
                -liq["retencion_boleta"], liq["liquido"], liq["sueldo_base"]
            ]
        else:
            vals = [
                liq["nombre"], liq.get("rut",""), liq.get("cargo",""),
                liq["tipo_contrato"].capitalize(), liq.get("afp",""), liq.get("salud",""),
                liq["sueldo_base"], b_col, b_mov,
                -liq["descuento_afp"], -liq["descuento_salud"],
                -liq["descuento_afc"], -liq["impuesto_unico"],
                -liq["total_descuentos"], liq["liquido"], liq["costo_empresa"]
            ]

        for j, v in enumerate(vals):
            c = ws.cell(row=rn, column=j+1, value=v)
            c.font = Font(name="Arial", size=9)
            c.fill = fill(bg)
            c.border = brd()
            if j >= 6 and isinstance(v, (int, float)):
                c.number_format = "$#,##0;($#,##0);\"-\""
                if v < 0:
                    c.font = Font(name="Arial", size=9, color="991B1B")
                elif j == 14:
                    c.font = Font(name="Arial", size=9, bold=True, color="166534")
                elif j == 15:
                    c.font = Font(name="Arial", size=9, bold=True, color="1E40AF")

    # Fila totales
    tr = len(data["liquidaciones"]) + 7
    ws.cell(row=tr, column=14, value="TOTALES").font = Font(name="Arial", bold=True, size=10)
    ws.cell(row=tr, column=14).fill   = fill(GRAY)
    ws.cell(row=tr, column=14).border = brd()
    for col, val in [(15, data["total_liquidos"]), (16, data["total_costo_empresa"])]:
        c = ws.cell(row=tr, column=col, value=val)
        c.font          = Font(name="Arial", bold=True, size=11, color=NAVY)
        c.number_format = "$#,##0;($#,##0);\"-\""
        c.fill          = fill(GRAY)
        c.border        = brd()

    # Anchos columnas
    widths = [22, 12, 16, 12, 10, 8, 12, 10, 10, 10, 10, 10, 10, 12, 12, 14]
    for i, w in enumerate(widths):
        ws.column_dimensions[chr(65+i)].width = w

    wb.save(buffer)
